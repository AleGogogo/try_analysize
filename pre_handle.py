import os
import pandas as pd
import re
import xlrd
import xlwt
from game import Game
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
def mergeToOne(sheet_name):
    excel = pd.read_excel('游戏top100.xls',sheet_name=sheet_name+'TOP100',header=0,encoding = 'utf-8')
    game_list = {}

    #按行遍历生成game对象
    for indexs,row in excel.iterrows():
        name = row['游戏名']
        downloadUrl=row['下载地址']
        versionName = row['版本号']
        file_md5 = row['MD5']
        pkName = row['包名']
        downloadCount = row['下载量(万)']
        game = Game(name, downloadUrl, downloadCount, file_md5, pkName, versionName)
        pattern = re.compile(u'[\u4e00-\u9fa5]+\d*|[a-zA-Z]+')
        c_name = re.findall(pattern, name)
        key = ''
        for c in c_name:
            key += str(c)
        if game_list.get(key) is not None:
            pre = game_list.get(key)
            if downloadCount >= pre.downloadCount:
               game_list.update({key: game})
        else:
            game_list.update({key: game})
    return game_list

def mergeRankInfo(fileName,gamelist):
    fileName = fileName + ".xlsx"
    workbook = load_workbook(fileName)
    sheet = workbook['Sheet1']  # index为0为第一张表
    row0 = ['排行', '游戏名', '下载地址', 'MD5', '包名', '版本号']  # 第一行内容
    for i in range(1, len(row0) + 1):
        if i == 1:
            continue
        if i == 2:
            continue
        else:
            sheet.cell(row=1, column=i).value = row0[i - 1]
            print(row0[i - 1])
    write_info(sheet,row0,gamelist)

    workbook.save(fileName)

def write_info(sheet,row0,game_list):
    i = 0
    for row in sheet['A1:F101']:
        i += 1
        print(i)
        for cell in row:
            name = cell.value
            print(name)
            if game_list.get(name) is not None:
                game = game_list.get(name)
                sheet.cell(row=i, column=3).value = game.downloadUrl
                sheet.cell(row=i, column=4).value = game.file_md5
                sheet.cell(row=i, column=5).value = game.pkName
                sheet.cell(row=i, column=6).value = game.versionName
                continue
            else:
                pass


